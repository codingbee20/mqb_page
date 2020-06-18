from django.shortcuts import render
from django.contrib.auth.models import Group, User
from mqbproductivity.models import TasksTracking, Daily, Effort, BlockData, TaskCategory
from management.models import Holiday, MonthlyProductivity, ActivitiesTask
from django.http import HttpResponseRedirect, JsonResponse
from django.urls import reverse
from urllib.parse import urlencode
import datetime
from datetime import datetime as datet
from openpyxl import Workbook
from django.http import StreamingHttpResponse, HttpResponse


# Create your views here.
def export_data(request):
    # check permission
    if request.method == 'POST':
        if 'user_list' in request.POST:
            file_name = datet.strftime(datet.now(), '%A_%d_%b_%Y_%H_%M_%S')
            month = request.POST.get('month')
            user_list = request.POST.getlist('user_list')
            selected_date = datet.strptime(month, '%B_%Y')
            start_date_Of_month = (selected_date.replace(day=1))
            end_date_Of_month = datetime.date(selected_date.year + selected_date.month // 12,
                                   selected_date.month % 12 + 1, 1) - datetime.timedelta(1)
            dates = Daily.objects.filter(date__range=[start_date_Of_month, end_date_Of_month])
            wb = Workbook()
            for user in user_list:
                user_id = User.objects.get(pk=user)
                ws = wb.create_sheet(user_id.username)
                ws.cell(1,1).value = "Date"
                ws.cell(1,2).value = "Categories"
                ws.cell(1,3).value = "Task_Name"
                ws.cell(1,4).value = "No.TC"
                ws.cell(1,5).value = "Effort"
                ws.cell(1,6).value = "Comments"
                row_num = 2
                for d in dates:
                    temp = Effort.objects.filter(dateCreateTask=d).filter(member=user_id)
                    for t in temp:
                        ws.cell(row_num,1).value = datet.strftime(d.date, '%A_%d_%b_%Y')
                        ws.cell(row_num,2).value = t.tasktracking.categories.name
                        ws.cell(row_num,3).value = t.tasktracking.taskName
                        ws.cell(row_num,4).value = t.quantity
                        ws.cell(row_num,5).value = t.effort
                        ws.cell(row_num,6).value = t.comment
                        row_num += 1
            std = wb.get_sheet_by_name('Sheet')
            wb.remove_sheet(std)
            filename = file_name+".xlsx"
            response = HttpResponse(content_type = 'application/vnd.ms-excel')
           # response['content_type'] = 'application/vnd.ms-excel'
            response['Content-Disposition'] = f'attachement; filename="{filename}"'
            wb.save(response)
            return response
            
    elif request.user.has_perm('auth.view_group'):
        if request.user.is_superuser:  # Check superuser can access to all groups
            groups = Group.objects.all()
        else:
            groups = request.user.groups.all()  # get user in group
        groups_dict = {}
        for group in groups:
            userslist = User.objects.filter(groups__name=group)
            listTemp = list()
            for user in userslist:
                listTemp.append({
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                })
            groups_dict[group.name] = listTemp
        contain = {
            'groups_dict': groups_dict,
        }
        return render(request, 'export_data.html', contain)

def productivity_management(request):
    if request.method == "POST":
        # change task status
        if 'queryDate' in request.POST:
            start_date = request.POST['startDate']
            end_date = request.POST['endDate']
            date_contain = {
                'startDate': start_date,
                'endDate': end_date,
            }
            base_url = reverse('productivity_management')
            query_string = urlencode(date_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

    date_today = datetime.date.today()
    if 'startDate' in request.GET and 'endDate' in request.GET:
        start_date = datet.strptime(request.GET.get('startDate'), '%Y-%m-%d')
        stop_date = datet.strptime(request.GET.get('endDate'), '%Y-%m-%d')
        group_selected = request.GET.get('group')
        member_selected = request.GET.get('member')
    else:
        start_date = date_today - datetime.timedelta(date_today.weekday())
        stop_date = start_date + datetime.timedelta(6)
        group_selected = None
        member_selected = None

    # check permission

    if request.user.has_perm('auth.view_group'):
        if request.user.is_superuser:  # Check superuser can access to all groups
            groups = Group.objects.all()
        else:
            groups = request.user.groups.all()  # get user in group
        groups_dict = {}
        efforts_dict = {}
        for group in groups:
            userslist = User.objects.filter(groups__name=group)
            listTemp = list()
            for user in userslist:
                listTemp.append({
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                })

                # query productivity

                dates = Daily.objects.filter(date__range=[start_date, stop_date])
                efforts = dict()
                for d in dates:
                    tasks = Effort.objects.filter(dateCreateTask=d).filter(member=user)
                    temp = list()
                    for e in tasks:
                        temp.append({
                            'effort_id': e.id,
                            'quantity': e.quantity,
                            'effort': e.effort,
                            'comment': e.comment,
                            'taskName': e.tasktracking.taskName,
                            'category': e.tasktracking.categories.name,
                        })
                    #print(temp)
                    efforts[datet.strftime(d.date, '%A_%d_%b_%Y')] = temp
                efforts_dict[user.username] = efforts
                # end query
            groups_dict[group.name] = listTemp

        categories = TaskCategory.objects.all()

        contain = {
            'groups_dict': groups_dict,
            'efforts_dict': efforts_dict,
            'start_date': start_date,
            'stop_date': stop_date,
            'group': group_selected,
            'member': member_selected,
            'categories': categories,
        }
    else:
        return render(request, "cannot_access_page.html", {})
    return render(request, 'productivity_management.html', contain)


def tasks_management(request):
    if request.is_ajax():
        task_id = request.POST.get('task_id')
        obj = TasksTracking.objects.filter(pk=int(task_id))
        if obj:
                if obj[0].modify_status:
                    obj.update(modify_status=False)
                elif obj[0].status:
                    obj.update(status=False, modify_status=False)
                else:
                    obj.update(status=True, modify_status=False)
        del obj
        obj = TasksTracking.objects.filter(pk=int(task_id))
        contain = {
            'task_id': obj[0].id,
            'status': obj[0].status,
            'modify_status': obj[0].modify_status,
        }
        print(contain)
        return JsonResponse(contain)
    
    if request.method == 'POST':
        # change task status
        if "change_task_status" in request.POST:
            taskid = request.POST['change_task_status']
            obj = TasksTracking.objects.filter(pk=int(taskid))

            if obj:
                if obj[0].modify_status:
                    obj.update(modify_status=False)
                elif obj[0].status:
                    obj.update(status=False, modify_status=False)
                else:
                    obj.update(status=True, modify_status=False)

            print(request.POST)
            selecteddate = datet.strptime(request.POST['selected_date'], '%B %Y')
            date_contain = {
                'month': selecteddate.strftime('%B_%Y'),
                'group': request.POST['group'],

            }
            base_url = reverse('tasks_management')
            query_string = urlencode(date_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

        if 'queryMonth' in request.POST:
            # startDate = datet.strptime(request.POST['startDate'], '%B %Y')

            selecteddate = request.POST['startDate'].replace(' ', '_')
            date_contain = {
                'month': selecteddate,
            }
            base_url = reverse('tasks_management')
            query_string = urlencode(date_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)
    
    if 'month' in request.GET:
        selectdate = datet.strptime(request.GET.get('month'), '%B_%Y')
        try:
            groupselected = request.GET.get('group')

        except Exception as e:
            groupselected = None

            print(e)
    else:
        selectdate = datetime.date.today()
        groupselected = None

    startdateOfmonth = (selectdate.replace(day=1))  # get start date of month
    enddateOfmonth = datetime.date(selectdate.year + selectdate.month // 12,
                                   selectdate.month % 12 + 1, 1) - datetime.timedelta(1)  # get end date of month
    enddateOfyear = (selectdate.replace(day=31, month=12))
    #next_three_month = startdateOfmonth + datetime.timedelta(6 * 365 / 12)
    # check permission
    print("end date of year")
    print(enddateOfyear)
    if request.user.has_perm('auth.view_group'):
        if request.user.is_superuser:  # Check superuser can access to all groups
            groups = Group.objects.all()
        else:
            groups = request.user.groups.all()

        groups_dict = {}
        tasks_tracking_dict = {}
        for group in groups:
            userslist = User.objects.filter(groups__name=group)
            listTemp = list()
            print(userslist)
            for user in userslist:
                listTemp.append({
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                })
                listTasks = TasksTracking.objects.filter(
                    stopdate__range=[startdateOfmonth, enddateOfyear]).filter(member=user)
                listTasksTemp = []
                for task in listTasks:
                    if task.startdate <= enddateOfmonth:
                        listTasksTemp.append({
                            'taskid': task.id,
                            'categories': task.categories.name,
                            'taskName': task.taskName,
                            'effort': task.effort_est,
                            'startdate': task.startdate,
                            'stopdate': task.stopdate,
                            'status': task.status,
                            'modify_status': task.modify_status
                        })
                tasks_tracking_dict[user.username] = listTasksTemp

            groups_dict[group.name] = listTemp
        contain = {
            'groups_dict': groups_dict,
            'tasks_tracking_dict': tasks_tracking_dict,
            'selectDate': startdateOfmonth,
            'group': groupselected,
        }
    else:
        return render(request, "cannot_access_page.html", {})
    return render(request, 'tasksmanagement.html', contain)


def baseline_effort(request):
    if request.method == 'POST':
        print(request.POST)
        if 'queryMonth' in request.POST:
            month_selected = request.POST['startDate'].replace(' ', '_')
            date_contain = {
                'month_selected': month_selected,
            }
            base_url = reverse('baseline')
            query_string = urlencode(date_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

        if 'baseline' in request.POST:
            month_selected = request.POST['month'].replace(' ', '_')
            group = request.POST['group']
            user_id = request.POST['user_id']
            user = User.objects.get(pk=user_id)

            selected_date = datet.strptime(month_selected, '%B_%Y')
            start_date_Of_month = (selected_date.replace(day=1))  # get start date of month

            Obj_Block_data = BlockData.objects.filter(member=user).filter(month=start_date_Of_month).first()
            try:
                if Obj_Block_data.status:
                    BlockData.objects.filter(member=user).filter(month=start_date_Of_month).update(status=False)
                else:
                    BlockData.objects.filter(member=user).filter(month=start_date_Of_month).update(status=True)
            except Exception as e:
                print(e)

            date_contain = {
                'month_selected': month_selected,
                'group': group,
            }
            base_url = reverse('baseline')
            query_string = urlencode(date_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

    if 'month_selected' in request.GET:
        try:
            groupselected = request.GET.get('group')
            selectdate = datet.strptime(request.GET.get('month_selected'), '%B_%Y')
        except Exception as e:
            groupselected = None
            selectdate = datetime.date.today().replace(day=1)
            print(e)
    else:
        selectdate = datetime.date.today().replace(day=1)
        groupselected = None

    if request.user.has_perm('auth.view_group'):
        if request.user.is_superuser:  # Check superuser can access to all groups
            groups = Group.objects.all()
        else:
            groups = request.user.groups.all()
        groups_dict = {}
        baseline_status = dict()
        for group in groups:
            userslist = User.objects.filter(groups__name=group)
            listTemp = list()
            for user in userslist:
                listTemp.append({
                    'id': user.id,
                    'username': user.username,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'email': user.email,
                })
                try:
                    date_user = BlockData.objects.filter(member=user).filter(month=selectdate).first()
                    baseline_status[user.username] = date_user.status
                except Exception as e:
                    baseline_status[user.username] = None
                    print(e)

            groups_dict[group.name] = listTemp
        contain = {
            'groups_dict': groups_dict,
            'selectDate': selectdate,
            'groupselected': groupselected,
            'baseline_status': baseline_status,
        }
    else:
        return render(request, "cannot_access_page.html", {})
    return render(request, 'baseline.html', contain)


def holiday(request):
    if request.method == 'POST':
        if 'add_date' in request.POST:
            date_selected = request.POST['holiday']
            update_Holiday = Holiday(
                date=date_selected,
            )
            update_Holiday.save()

            selectedYear = datet.strptime(date_selected, '%Y-%m-%d').year
            contain = {
                'selectedYear': selectedYear,
            }
            base_url = reverse('holiday')
            query_string = urlencode(contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

        if 'delete_date' in request.POST:
            Holiday.objects.filter(id=request.POST['Date']).delete()
            selectedYear = request.POST['selectedYear']
            contain = {
                'selectedYear': selectedYear,
            }
            base_url = reverse('holiday')
            query_string = urlencode(contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

    if 'query_year' in request.POST:
        selected_year = int(request.POST['selectedYear'])
    elif 'selectedYear' in request.GET:
        selected_year = int(request.GET.get('selectedYear'))
    else:
        selected_year = datetime.date.today().year

    print(selected_year)
    start_day = datetime.date(selected_year, 1, 1)
    end_date = datetime.date(selected_year, 12, 31)
    holiday_day = Holiday.objects.filter(date__range=[start_day, end_date])

    year_list = list()
    for y in range(2019, 2025):
        year_list.append(y)

    contain = {
        'holiday': holiday_day,
        'year_list': year_list,
        'selected_year': selected_year,

    }
    return render(request, 'holiday.html', contain)


def monthly_productivity(request):
    if request.user.is_superuser:  # Check superuser can access to all groups
        groups = Group.objects.all()
    else:
        groups = request.user.groups.all()

    if request.method == 'POST':
        if 'query_year' in request.POST:
            selected_year = request.POST['selectedYear']
            year_contain = {
                'year': selected_year,
            }
            base_url = reverse('monthly_productivity')
            query_string = urlencode(year_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

        if 'update_db' in request.POST:
            year = int(request.POST['selectedYear'])
            num_date_working = 0
            start_day_of_year = datetime.date(year, 1, 1)
            end_date_of_year = datetime.date(year, 12, 31)
            holidays = Holiday.objects.filter(date__range=[start_day_of_year, end_date_of_year])
            holidays_list = list()
            for h in holidays:
                holidays_list.append(h.date)
            if groups:
                for group in groups:
                    users = User.objects.filter(groups__name=group)

                    # loop for 12 months in year
                    for current_month in range(1, 13):
                        # get start/end date of month
                        startdateOfmonth = datetime.date(year, current_month, 1)
                        enddateOfmonth = datetime.date(startdateOfmonth.year + startdateOfmonth.month // 12,
                                                       startdateOfmonth.month % 12 + 1, 1) - datetime.timedelta(1)
                        # --------end----------

                        # get number of working date and number of working hour
                        date_in_month = startdateOfmonth
                        while date_in_month <= enddateOfmonth:
                            if date_in_month not in holidays_list and date_in_month.strftime("%A") not in ['Sunday',
                                                                                                           'Saturday']:
                                num_date_working += 1
                            date_in_month += datetime.timedelta(1)

                        num_hour_working = num_date_working * 8
                        # --------end---------

                        dates = Daily.objects.filter(date__range=[startdateOfmonth, enddateOfmonth])
                        tasks = list()
                        for d in dates:
                            tasks.append(Effort.objects.filter(dateCreateTask=d).filter(member__in=users))

                        task_list = list(TaskCategory.objects.all())
                        for t in task_list:
                            totaltc = 0
                            totale = 0
                            for eff in tasks:
                                for e in eff:
                                    if e.tasktracking.categories.name == t.name:
                                        totaltc += e.quantity
                                        totale += e.effort

                            # update info for PyS
                            if totale / num_hour_working != 0:
                                pys_result = totaltc / (totale / num_hour_working) / num_date_working
                            else:
                                pys_result = 0.0

                            # update PyS to database
                            # update ActivityTask
                            try:
                                atc = ActivitiesTask.objects.filter(group_name=group).filter(task_name=t.name)
                                if not atc:
                                    ActivitiesTask(
                                    group_name=group,
                                    task_name=t.name,
                                    ).save()
                            except Exception as e:
                                print(e)

                            # try:
                            #     ActivitiesTask(
                            #         group_name=group,
                            #         task_name=t.name,
                            #     ).save()
                            # except Exception as e:
                            #     print(e)
                            # update MonthlyProductivity
                            update_activity_task = ActivitiesTask.objects.filter(group_name=group).filter(task_name=t.name).first()
                            if MonthlyProductivity.objects.filter(activity_task=update_activity_task).filter(year=year).filter(month=current_month):
                                MonthlyProductivity.objects.filter(activity_task=update_activity_task).filter(year=year).filter(month=current_month).update(
                                    pys=pys_result,
                                )
                            else:
                                MonthlyProductivity(
                                    activity_task=update_activity_task,
                                    year=year,
                                    month=current_month,
                                    pys=pys_result,
                                ).save()
                            # --------end - update PyS to database------

                year_contain = {
                    'year': year,
                }
                base_url = reverse('monthly_productivity')
                query_string = urlencode(year_contain)
                url = '{}?{}'.format(base_url, query_string)
                return HttpResponseRedirect(url)
            else:
                # if user doesn't belong to any group
                contain = {
                    'notification': 'You are not in any group. Please contact Administrator!!!'
                }
                return render(request, 'monthly_productivity.html', contain)

        if 'add_target' in request.POST:
            data = dict(request.POST)
            print(data)
            for i, value in enumerate(data['activity_task']):
                year = data['selectedYear'][i]
                current_month = data['selectedMonth'][i]
                pys_result = data['pys'][i]
                try:
                    group_user = Group.objects.get(name=data['group_user'][i])
                    if not ActivitiesTask.objects.filter(group_name=group_user).filter(task_name=data['activity_task'][i]):
                        ActivitiesTask(
                            group_name=group_user,
                            task_name=data['activity_task'][i],
                        ).save()
                except Exception as e:
                    print(e)
                # update MonthlyProductivity
                update_activity_task = ActivitiesTask.objects.filter(task_name=data['activity_task'][i]).first()
                if MonthlyProductivity.objects.filter(activity_task=update_activity_task).filter(year=year).filter(
                        month=current_month):
                    MonthlyProductivity.objects.filter(activity_task=update_activity_task).filter(year=year).filter(
                        month=current_month).update(
                        pys=pys_result,
                    )
                else:
                    MonthlyProductivity(
                        activity_task=update_activity_task,
                        year=year,
                        month=current_month,
                        pys=pys_result,
                    ).save()

            year_selected = request.POST['selected_year']
            year_contain = {
                'year': year_selected,
            }
            base_url = reverse('monthly_productivity')
            query_string = urlencode(year_contain)
            url = '{}?{}'.format(base_url, query_string)
            return HttpResponseRedirect(url)

    if 'year' in request.GET:
        year = int(request.GET.get('year'))
    else:
        year = datetime.date.today().year

    # get working date in month base on Holiday list.
    # Saturday and Sunday are not working date
    monthly_productivity_chart = {}
    
    month_dict = {
        1: 'Jan',
        2: 'Feb',
        3: 'Mar',
        4: 'Apr',
        5: 'May',
        6: 'Jun',
        7: 'Jul',
        8: 'Aug',
        9: 'Sep',
        10: 'Oct',
        11: 'Nov',
        12: 'Dec',
    }
    if groups:
        for group in groups:
            yearly_productivity = [['Month',
                            'Working TI', {'type': 'number', 'role': 'annotation'},
                            'Completed TI', {'type': 'number', 'role': 'annotation'},
                           'Reviewed TI',  {'type': 'number', 'role': 'annotation'},
                            'Completed TI Target',
                            'Reviewed TI Target', {'type': 'number', 'role': 'annotation'}
                            ]]
            try:
                working_ti_activities_task = ActivitiesTask.objects.filter(group_name=group).filter(
                    task_name="Working_TI").first()
                ti_activities_task = ActivitiesTask.objects.filter(group_name=group).filter(
                    task_name="Test_Auto").first()
                re_ti_activities_task = ActivitiesTask.objects.filter(group_name=group).filter(
                    task_name="Test_Auto_Review").first()
                completed_ti_target = ActivitiesTask.objects.filter(group_name=group).filter(
                    task_name="Completed_TI_Target").first()
                reviewed_ti_target = ActivitiesTask.objects.filter(group_name=group).filter(
                    task_name="Reviewed_TI_Target").first()
            except Exception as e:
                print(e)

            for m in range(1, 13):
                if year == 2019 and (m in [1, 2, 3, 4, 5, 6, 7, 8]) and group.name == 'MQB':
                    if m == 1:
                        continue
                        yearly_productivity.append([month_dict[m], 0, 0, 0, 0, 0, 0, 5, 25])
                    if m == 2:
                        continue
                        yearly_productivity.append([month_dict[m], 0, 0, 0, 0, 0, 0, 5, 25])
                    if m == 3:
                        yearly_productivity.append([month_dict[m], 0, 0, 1.66, 1.66, 2.96, 2.96, 5, 25, 25])
                    if m == 4:
                        yearly_productivity.append([month_dict[m], 0, 0, 3.23, 3.23, 8.98, 8.98, 5, 25, 25])
                    if m == 5:
                        yearly_productivity.append([month_dict[m], 0, 0, 2.48, 2.48, 3.15, 3.15, 5, 25, 25])
                    if m == 6:
                        yearly_productivity.append([month_dict[m], 0, 0, 1.34, 1.34, 7.33, 7.33, 5, 25, 25])
                    if m == 7:
                        yearly_productivity.append([month_dict[m], 0, 0, 1.35, 1.35, 15.22, 15.22, 5, 25, 25])
                    if m == 8:
                        yearly_productivity.append([month_dict[m], 5.18, 5.18, 1.88, 1.88, 11.73, 11.73, 5, 25, 25])
                else:
                    temp = [month_dict[m]]
                    try:
                        working_ti_monthly_productivity = MonthlyProductivity.objects.filter(
                            activity_task=working_ti_activities_task).filter(year=year).get(month=m)
                        temp.append(round(working_ti_monthly_productivity.pys, 2))
                        temp.append(round(working_ti_monthly_productivity.pys, 2))
                    except Exception as e:
                        print(e)
                        temp.append(0)
                        temp.append(0)
                    try:
                        ti_monthly_productivity = MonthlyProductivity.objects.filter(
                            activity_task=ti_activities_task).filter(year=year).get(month=m)
                        temp.append(round(ti_monthly_productivity.pys, 2))
                        temp.append(round(ti_monthly_productivity.pys, 2))
                    except Exception as e:
                        print(e)
                        temp.append(0)
                        temp.append(0)
                    try:
                        re_ti_monthly_productivity = MonthlyProductivity.objects.filter(
                            activity_task=re_ti_activities_task).filter(year=year).get(month=m)
                        temp.append(round(re_ti_monthly_productivity.pys, 2))
                        temp.append(round(re_ti_monthly_productivity.pys, 2))
                    except Exception as e:
                        print(e)
                        temp.append(0)
                        temp.append(0)
                    try:
                        completed_ti_target_monthly_productivity = MonthlyProductivity.objects.filter(
                            activity_task=completed_ti_target).filter(year=year).get(month=m)
                        temp.append(round(completed_ti_target_monthly_productivity.pys, 2))
                    except Exception as e:
                        print(e)
                        temp.append(0)
                    try:
                        reviewed_ti_target_monthly_productivity = MonthlyProductivity.objects.filter(
                            activity_task=reviewed_ti_target).filter(year=year).get(month=m)
                        temp.append(round(reviewed_ti_target_monthly_productivity.pys, 2))
                        temp.append(round(reviewed_ti_target_monthly_productivity.pys, 2))
                    except Exception as e:
                        print(e)
                        temp.append(0)
                        temp.append(0)

                    # Completed TI and Review TI are not 0 in month
                    print(temp)
                    if temp[2] != 0 or temp[3] != 0 or len(temp)>0:
                        yearly_productivity.append(temp)

            monthly_productivity_chart[group.name] = yearly_productivity
    print(monthly_productivity_chart)
    year_list = list()
    for y in range(2019, 2025):
        year_list.append(y)
    contain = {
        'monthly_productivity_chart': monthly_productivity_chart,
        'year_list': year_list,
        'selected_year': year,
    }
    return render(request, 'monthly_productivity.html', contain)
