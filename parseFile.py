from openpyxl import load_workbook
import re
import datetime

class EXCEL_Pro:
	
	def __init__(self, path):
		self.path = path
		self.bigData = {}
		self.wb = load_workbook(self.path, data_only=True)
	
	####Get the list name of worksheet#######
		self.Get_Data()
		
	
	def Get_Data(self):
		user_list = self.wb.sheetnames
		print(user_list)
		for user in user_list:
			date_col = list()
			date_list = list()
			for rows in self.wb[user].rows:
				for i, cell in enumerate(rows):
					print(cell.value)
					try:
						date_list.append(cell.value.date())
						date_col.append(i+1)
					except Exception as e:
						pass
				break
			print(date_col)
			print(date_list)
			break
if __name__== '__main__':
	
	path = "C:\\UYO1HC\\mqb_page\\importdata\\static\\Document_effort.xlsx"
	data = EXCEL_Pro(path)