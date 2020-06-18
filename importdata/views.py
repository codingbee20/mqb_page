from django.shortcuts import render
from mqbproductivity.models import TaskCategory, Daily, BlockData, TasksTracking, Effort
from django.contrib.auth.models import User, Group
# Create your views here.
from openpyxl import load_workbook
import datetime


def add_effort(user_id, date_create_task, task_obj, quantity, effort, comment=''):
    update_effort = Effort(
            member=user_id,
            dateCreateTask=date_create_task,
            tasktracking=task_obj,
            quantity=quantity,
            effort=effort,
            comment=comment,
    )
    update_effort.save()
    return update_effort


def import_data(request):

    path = "importdata\\static\\Document_effort_Dec.xlsx"
    wb = load_workbook(path, data_only=True)
    print(wb.sheetnames)
    users_list = wb.sheetnames
    # create user
    if not Group.objects.get(name='MQB'):
        mqb_group = Group.objects.create(name='MQB')
    else:
        mqb_group = Group.objects.get(name='MQB')

    for user in users_list:
        if not User.objects.filter(username=user):
            user_id = User.objects.create_user(username=user, email=user+"@vn.bosch.com", password="12345678q@Q")
        else:
            user_id = User.objects.get(username=user)
        mqb_group.user_set.add(user_id)
        task_tracking = [
            'Test_Spec',
            'Test_Spec_Review',
            'Test_Auto',
            'Test_Auto_Review',
            'SW_execution',
            'Project_task_activities',
            'Project_misc',
            'Non-Project'
        ]
        # create task
        task_row = list()
        task_id = list()
        for t in task_tracking:
            category = TaskCategory.objects.get(name=t)
            update_task = TasksTracking(
                member=user_id,
                categories=category,
                taskName=category.name,
                effort_est=168,
                startdate="2019-12-01",
                stopdate="2019-12-31",
                status=False,
                modify_status=False,
            )
            update_task.save()
            for r in wb[user].rows:
                if t == r[0].value:
                    task_row.append(r[0].row)
                    task_id.append(update_task.id)
        print(task_row)
        print(task_id)
        date_col = list()
        date_list = list()

        for row in wb[user].rows:
            for i, cell in enumerate(row):
                try:
                    date_list.append(cell.value.date())
                    date_col.append(i + 1)
                except Exception as e:
                    pass
            break

        # add task
        for j, d in enumerate(date_list):
            if Daily.objects.filter(date=d):
                date_input_task = Daily.objects.get(date=d)
            else:
                date_input_task = Daily(date=d)
                date_input_task.save()
            for i, t in enumerate(task_id):
                task_input = TasksTracking.objects.get(id=t)
                quantity = wb[user].cell(task_row[i], date_col[j]).value
                effort = wb[user].cell(task_row[i], date_col[j]+1).value
                if quantity != 0 or effort != 0:
                    add_effort(user_id, date_input_task, task_input, quantity, effort, "")

        # add block_data
        if not BlockData.objects.filter(month="2019-12-01").filter(member=user_id):
            update_BlockData = BlockData(
                month="2019-12-01",
                member=user_id,
                status=False
            )
            update_BlockData.save()
            del update_BlockData
    contain = {
        'status': 'Successful!!!!'
    }
    return render(request, "IO_ports.html", contain)
